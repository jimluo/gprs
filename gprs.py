#!/usr/bin/env python
#coding:utf8
import sys
import email
import poplib
import json
from openpyxl import Workbook
from openpyxl.cell import Cell
# from openpyxl.style import Border
# from openpyxl.chart import ScatterChart, Serie, Reference
# from openpyxl.style import Color
import datetime
import sqlite3
import matplotlib as mp
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.font_manager import fontManager, FontProperties
from PyQt5 import QtCore, QtWidgets, Qt, QtGui
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem
from PyQt5.QtCore import QUrl
from ui_gprs import Ui_gprs
# from ui_config import Ui_config
# from ui_help import Ui_help

mp.rcParams['font.sans-serif'] = ['SimHei']

class Help(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(Help, self).__init__(parent)
        self.ui = ui = Ui_help()
        ui.setupUi(self)

class Config(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(Config, self).__init__(parent)
        self.ui = ui = Ui_config()
        ui.setupUi(self)
        self.cfg = cfg = json.load(open('config.txt', 'rt'))

        ui.mail.setText(cfg['mail'])
        ui.name.setText(cfg['name'])
        ui.pwd.setText(cfg['pwd'])
        ui.current.setValue(cfg['current'])

        wells = ui.wells
        wells.setColumnWidth(3, 80)
        i = 0
        wells.insertRow(i)
        for f, groups in cfg['field'].iteritems():
            for g, sites in groups.iteritems():
                for n, site in enumerate(sites):
                    wells.setItem(i, 0, QTableWidgetItem(f))
                    wells.setItem(i, 1, QTableWidgetItem(g))
                    for j, s in enumerate(site):
                        wells.setItem(i, j+2, QTableWidgetItem(str(s)))
                    # if n < len(sites)-1:
                    i = i + 1
                    wells.insertRow(i)
        wells.removeRow(i)

    def saveCfg(self):
        ui = self.ui
        self.cfg['mail'] = str(ui.mail.text())
        self.cfg['name'] = str(ui.name.text())
        self.cfg['pwd'] =  str(ui.pwd.text())
        self.cfg['current'] = ui.current.value()
        self.cfg['field'] = {}
        wells = ui.wells
        for r in range(wells.rowCount()):
            row = []
            f = str(wells.item(r, 0).text().toUtf8())
            g = str(wells.item(r, 1).text())
            if not f in self.cfg['field']:
                self.cfg['field'][f] = {}
            if not g in self.cfg['field'][f]:
                self.cfg['field'][f][g] = []
            for c in range(2,10):
                row.append(str(wells.item(r, c).text()))
            self.cfg['field'][f][g].append(row)

        f=open('config.txt', 'w')
        f.write(json.dumps(self.cfg, sort_keys=True, indent=4))
        f.close()

        self.close()

    def appendRow(self):
        self.ui.wells.insertRow(self.ui.wells.currentRow())

    def deleteRow(self):
        self.ui.wells.removeRow(self.ui.wells.currentRow())

class Gprs(QtWidgets.QMainWindow):
    #初始化
    def __init__(self, parent=None):
        super(Gprs, self).__init__(parent)
        self.ui = ui = Ui_gprs()
        self.ui.setupUi(self)

        ui.toolBar.insertWidget(ui.actionSample, ui.lblLogo);
        ui.toolBar.insertSeparator(ui.actionSample);
        ui.toolBar.insertWidget(ui.actionSample, ui.lblDate);
        ui.toolBar.insertWidget(ui.actionSample, ui.date);
        ui.toolBar.insertSeparator(ui.actionSample);

        self.alert = []

        for i, w in enumerate((50,80,120,80,80,80,80,80,80,80,80,80)):
            ui.wells.setColumnWidth(i, w)
        for i in range(150):
            ui.wells.insertRow(i)
        
        ui.date.setDate(datetime.date.today())
        self.conn = sqlite3.connect('puyuan.db')
        
        self.cfg = json.load(open('config.txt', 'rt'))
        self.phonesite, self.siteitem = {}, {}
        #读取config中数据填充树形列表
        for f, groups in self.cfg['field'].items():
            field = QtWidgets.QTreeWidgetItem((f,)) #添加井区
            ui.field.addTopLevelItem(field)
            for g, sites in groups.items():
                group = QtWidgets.QTreeWidgetItem((g,)) #添加井场
                field.addChild(group)
                for site in sites:
                    group.addChild(QtWidgets.QTreeWidgetItem((site[0],))) #添加井场
                    self.phonesite[site[1]] = site[0] #保存所有井厂相关手机号
                    self.siteitem[site[0]] = site[2:]

        ui.field.expandAll()

        #创建曲线图
        self.fig = fig = Figure(figsize=(ui.frame.width()/6, ui.frame.height()/6.5),dpi=100)
        fig.subplots_adjust(left=0.04, bottom=0.35, right=0.995, top=0.96)
        self.canvas = FigureCanvas(fig)
        FigureCanvas.setSizePolicy(self.canvas, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.canvas.setParent(ui.frame)
        self.fig.add_subplot(111)

        self.showMaximized()
        ui.split.setSizes((400,400))

    def on_btnHelp(self):
        help = Help(self)
        help.show()

    def on_btnConfig(self):
        config = Config(self)
        config.show()
        self.cfg = json.load(open('config.txt', 'rt'))

    #属性列表和日期控件选择出发事件
    def on_field_date_changed(self):
        wells = self.ui.wells
        site = self.ui.field.currentItem()
        if site != None and site.childCount() == 0:
            #按照当前选择日期和井名查询数据库
            sitename = str(site.text(0))
            date = self.ui.date.date().toString('yyyy.MM.')
            sql = "select * from wells where site='%s' and date like '%s'" % (sitename, date+'%')
            c = self.conn.cursor()
            rows = c.execute(sql)
            self.sites = sites = [row for row in rows]
            c.close()
            wells.clear()
            self.fig.clf()
            axe = self.fig.add_subplot(111)
            axe.grid(True)
            if len(sites) == 0:
                self.canvas.draw()
                return

            title = ['序号','井组','日期时间','总电压(V)','总电流(A)','回路电阻(Ω)'] + self.siteitem[sitename]
            wells.setHorizontalHeaderLabels([t for t in title])

            for i, row in enumerate(sites):
                wells.insertRow(i)
                wells.setItem(i, 0, QTableWidgetItem(str(i))) #电阻
                for j in range(4):
                    wells.setItem(i, j+1, QTableWidgetItem(str((row[j+2]))))
                wells.setItem(i, 5, QTableWidgetItem("%.1f" % (row[4]/(row[5]+0.0001)))) #电阻
                #6个电流
                for j, w in enumerate(self.siteitem[sitename]):
                    if w != '': #存在这口井，则显示电流
                        t = QTableWidgetItem('%2.2f' % row[6+j])
                        if 2 > row[6+j] or row[6+j] > 6:
                            t.setForeground(QtGui.QBrush(Qt.QColor('red')))
                        wells.setItem(i, 6+j, t)

            for i in range(len(sites), wells.rowCount()):
                wells.setRowHidden(i, True)

            x = range(len(sites)+1)
            for i, w in enumerate(self.siteitem[sitename]):
                if w != '':
                    current = [sites[0][6]] + [s[6+i] for s in sites]
                    axe.plot(x, current, color='rgbykc'[i], label='%s' % w)

            for labeltick in axe.xaxis.get_majorticklabels():
                labeltick.set_fontsize(8)
            axe.set_xticks(range(len(sites)+2))
            axe.set_xticklabels([s[3] for s in sites], rotation=45)
            axe.set_yticks(range(5-self.cfg['current'], 5+self.cfg['current']))
            axe.legend(title='井')
            axe.legend(prop=FontProperties(size='xx-small'))
            axe.set_xlabel('日期时间(年.月.日.时)')
            axe.set_ylabel('电流(A)')

            self.canvas.draw()

    #定时读取邮箱数据到数据库
    def on_btnMail(self):
        #登录邮箱
        p = poplib.POP3(self.cfg['mail'], 110)
        # p.set_debuglevel(1)
        try:
            p.user(self.cfg['name'])
            p.pass_(self.cfg['pwd'])
        except poplib.error_proto:
            QMessageBox.warning(self, '错误', '登陆139邮箱错误\n%s,%s' % (poplib.error_proto, e))
            p.quit()
            return

	    #下载每一封邮件，解析发件人、日期、主题
        c = self.conn.cursor()

        self.alert = []
        numrow = 0
        num,total_size = p.stat()
        for i in range(1, num+1):
            status = p.list(i)
            length = int(status.split()[-1])
            if length < 6000:
                mail = email.message_from_string('\n'.join(p.retr(i)[1]))
                phone, subject, date = mail['From'][:11], mail['Subject'], mail['Date']
                if phone.startswith('1') and len(phone) == 11: #手机号码长度
                    if phone in self.phonesite:
                        site = self.phonesite[phone]
                        date = '%4d.%02d.%02d.%02d' % email.Utils.parsedate_tz(date)[:4]
                        subject = email.Header.decode_header(subject)[0][0]
                        if len(subject) == 51:
                            subject = subject[3:]
                        currents = subject.split(',')
                        if len(subject) == 48:
                            floatcurrents = [0,0,0,0,0,0,0,0]
                            for i,current in enumerate(currents):
                                floatcurrents[i] = float(current)
                            if floatcurrents[0] < 1.1 or floatcurrents[1] < 0.1:
                                continue
                            row = [site+date, int(phone), site, date] + floatcurrents
                            try:
                                c.executemany('insert into wells values(?,?,?,?, ?,?, ?,?,?,?,?,?)', (row,))
                            except sqlite3.Error:
                                # self.statusBar().showMessage('insert DB error: %s, num:%d, idx:%d, len:%d. %s, %d, %s' % (e.args[0], num, i, length, site, phone, date))
                                continue

                            numrow += 1
                            self.statusBar().showMessage('下载数据:%s条,状态:%s条,请10秒钟后再继续下载和报表操作' % (numrow, len(self.alert)))
                        elif len(subject) == 1:
                            if currents[0] == '0' or currents[0] == '1':
                                state = '运行' if currents[0] == '1' else '停机'
                                self.alert.append('%s, %s, %s' % (site, date, state))

        self.ui.alert.clear()
        for a in self.alert:
            self.ui.alert.addItem(a)

        self.conn.commit()
        c.close()
        p.quit()

    #报表输出
    def on_btnReport(self):
        print('report')
        fname = QtGui.QFileDialog.getSaveFileName(self, '报表名称','', "Excell Files (*.xlsx)")
        if not fname:
            return

        site = self.ui.field.currentItem()
        sitename, fieldname = str(site.text(0)), site.parent().text(0).toUtf8()
        wb = Workbook()
        ws = wb.worksheets[0]
        cell = ws.cell(row=1, column=3)
        cell.style.font.size = 24
        cell.style.font.bold = True
        cell.value = "xx油田第三采油厂阴极保护系统数据表(%s)" % fieldname
        title = ['序号','井组','日期时间','总电压(V)','总电流(A)','回路电阻(Ω)'] + self.siteitem[sitename]
        title = [t for t in title if t != '']
        wells = self.ui.wells
        def format_cell(cell):
            cell.style.borders.top.border_style  = Border.BORDER_THIN
            cell.style.borders.bottom.border_style  = Border.BORDER_THIN
            cell.style.borders.left.border_style  = Border.BORDER_THIN
            cell.style.borders.right.border_style  = Border.BORDER_THIN
            cell.style.alignment.horizontal = 'center'
            return cell

        for j, t in enumerate(title): #header
            cell = format_cell(ws.cell(row=4, column=j))
            cell.value = t
        rownum = len(self.sites)
        for i in range(rownum):
            for j in range(len(title)):
                cell = format_cell(ws.cell(row=5+i, column=j))
                item = wells.item(i,j)
                if item != None:
                    cell.value = str(item.text())
                    if j > 5:
                        if 2 > float(item.text()) or float(item.text()) > 6:
                            # print i,j,float(item.text())
                            cell.style.font.color.index = Color.RED
                            cell.set_value_explicit(str(item.text()), Cell.TYPE_NUMERIC)
                            cell.style.number_format.format_code = '0.00'
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["F"].width = 12
        wb.save(str(fname))

    def on_webmail(self):
        QDesktopServices.openUrl(QUrl("http://mail.10086.cn/", QUrl.TolerantMode))

    def on_exit(self):
        self.conn.close()
        self.close()

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    gprs = Gprs()
    gprs.show()
    sys.exit(app.exec_())
